# -*- coding: utf-8 -*-
"""
Compila il file model-portfolios_compositions-template.xlsx
con i dati del portafoglio — solo colonne A (Data) e B (ISIN).
Preserva tutto il resto del template originale.
"""
import io
import datetime
from pathlib import Path
import openpyxl

_TEMPLATE = Path(__file__).parent.parent / "data" / "model_portfolio_template.xlsx"


def export_advisorelite_csv(funds: list[dict]) -> bytes:
    """
    Genera CSV formato AdvisorElite:
      ISIN,Amount,
      LU...,30
    Pesi interi normalizzati a 100.
    """
    total = sum(float(f.get("peso", 0)) for f in funds) or 1
    lines = ["ISIN,Amount,"]
    for f in sorted(funds, key=lambda x: x.get("bucket","") + x.get("ISIN","")):
        isin = str(f.get("ISIN","")).strip()
        peso = round(float(f.get("peso", 0)) / total * 100)
        lines.append(f"{isin},{peso}")
    return "\n".join(lines).encode("utf-8")


def export_portfolio_excel(funds: list[dict], portfolio_name: str = "") -> bytes:
    """
    Apre il template e compila le colonne:
      A = Data (DD/MM/YYYY)
      B = ISIN
    Lascia invariate tutte le altre colonne e la formattazione del template.
    """
    today = datetime.date.today().strftime("%d/%m/%Y")

    # Ordina per bucket → peso decrescente
    funds_sorted = sorted(funds, key=lambda x: (x.get("bucket",""), -x.get("peso", 0)))

    # Carica template se disponibile, altrimenti crea struttura minima
    if _TEMPLATE.exists():
        wb = openpyxl.load_workbook(str(_TEMPLATE))
        ws = wb["Compositions"] if "Compositions" in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Compositions"
        # Header minimale se manca il template
        ws["A1"] = "Composition date"
        ws["B1"] = "ISIN"

    # Trova la prima riga dati libera (dopo gli header)
    # Il template ha 2 righe di header (riga 1=descrizioni, riga 2=nomi colonne)
    first_data_row = 3

    # Scrivi Data e ISIN — solo colonne A e B
    for i, f in enumerate(funds_sorted):
        row = first_data_row + i
        isin = str(f.get("ISIN", "")).strip()

        # Colonna A: Data
        ws.cell(row=row, column=1, value=today)
        # Colonna B: ISIN
        ws.cell(row=row, column=2, value=isin)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
