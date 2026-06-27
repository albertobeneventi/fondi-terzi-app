# -*- coding: utf-8 -*-
"""
Aggiornamento UNIFICATO performance + volatilita 1Y da fondidoc.it.
Scarica UNA volta sola e aggiorna entrambi i file (stessi ISIN/link FondiDoc):

  data/fondi.xlsx          -> usato dall'app Streamlit
      header riga 1, dati da riga 2; ISIN col3; link FondiDoc col30;
      perf col21=1Y,22=3Y,23=YTD,24=2024,25=2023,26=2022,27=VOL
      schede 'tutti quelli trasferibili' (+ 'quelli gestibili' duplicato)

  data/tabella_fondi.xlsx  -> file di lavoro (formato "tabella arricchita")
      header riga 2, dati da riga 3; ISIN col3; link FondiDoc col23;
      perf col14=1Y,15=3Y,16=YTD,17=2024,18=2023,19=2022,20=VOL
      scheda 'tutti quelli trasferibili'

Gira in GitHub Actions (il 2 di ogni mese). Streamlit Cloud ridistribuisce
automaticamente l'app al commit.
"""
import re, time, datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import requests
from bs4 import BeautifulSoup
import openpyxl

BASE = Path(__file__).parent / "data"
WORKERS = 12
PERC_RE = re.compile(r'-?\d+[,.]\d+%')

# configurazione per ciascun file: schema colonne, riga iniziale dati, schede
FILES = [
    {
        "path": BASE / "fondi.xlsx",
        "data_start": 2, "col_isin": 3, "col_fondidoc": 30,
        "perf": {21: 'p1y', 22: 'p3y', 23: 'ytd', 24: 'p2024', 25: 'p2023', 26: 'p2022', 27: 'vol1y'},
        "sheets": ["tutti quelli trasferibili", "quelli gestibili"],
    },
    {
        "path": BASE / "tabella_fondi.xlsx",
        "data_start": 3, "col_isin": 3, "col_fondidoc": 23,
        "perf": {14: 'p1y', 15: 'p3y', 16: 'ytd', 17: 'p2024', 18: 'p2023', 19: 'p2022', 20: 'vol1y'},
        "sheets": ["tutti quelli trasferibili"],
    },
]

_tl = threading.local()


def log(m):
    print(f"[{datetime.datetime.now():%H:%M:%S}] {m}", flush=True)


def parse_pct(s):
    try:
        return round(float(s.replace('%', '').replace(',', '.')) / 100, 6)
    except Exception:
        return None


def get_session():
    s = getattr(_tl, "s", None)
    if s is None:
        s = requests.Session()
        s.headers.update({"User-Agent": "Mozilla/5.0"})
        _tl.s = s
    return s


def fetch(url, timeout=25):
    last = None
    for _ in range(3):
        try:
            return get_session().get(url, timeout=timeout)
        except Exception as e:
            last = e
            time.sleep(2)
    raise last


def scrape_fund(fd_url):
    result = {}
    r = fetch(fd_url)
    lines = [l.strip() for l in BeautifulSoup(r.text, "html.parser").get_text(separator="\n").split("\n") if l.strip()]
    for j, line in enumerate(lines):
        nxt = lines[j + 1] if j + 1 < len(lines) else ''
        if 'YTD' in line and ('1 anno' in line or '1Y' in line or 'anno' in nxt):
            vals = PERC_RE.findall(line) or (PERC_RE.findall(lines[j + 1]) if j + 1 < len(lines) else [])
            if vals:
                result['ytd'] = parse_pct(vals[0])
                result['p1y'] = parse_pct(vals[1]) if len(vals) > 1 else None
                result['p3y'] = parse_pct(vals[2]) if len(vals) > 2 else None
            break
    for j, line in enumerate(lines):
        if '2024' in line and '2023' in line and '2022' in line:
            vals = PERC_RE.findall(line) or (PERC_RE.findall(lines[j + 1]) if j + 1 < len(lines) else [])
            if len(vals) >= 3:
                if '2022' in line.split('2023')[0]:
                    result['p2022'], result['p2023'], result['p2024'] = parse_pct(vals[0]), parse_pct(vals[1]), parse_pct(vals[2])
                else:
                    result['p2024'], result['p2023'], result['p2022'] = parse_pct(vals[0]), parse_pct(vals[1]), parse_pct(vals[2])
            break
    try:
        r2 = fetch(fd_url.replace("/d/Index/", "/d/Ana/"))
        lines2 = [l.strip() for l in BeautifulSoup(r2.text, "html.parser").get_text(separator="\n").split("\n") if l.strip()]
        for j, line in enumerate(lines2):
            if 'olatil' in line and 'negativa' not in line:
                for k in range(j + 1, min(j + 8, len(lines2))):
                    if PERC_RE.match(lines2[k]):
                        result['vol1y'] = parse_pct(lines2[k])
                        break
                break
    except Exception:
        pass
    return result


def fondidoc_url(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    v = cell.value
    return v if isinstance(v, str) and v.startswith('http') else None


def collect_isin_urls():
    """Raccoglie ISIN->URL FondiDoc unendo entrambi i file (dedup per ISIN)."""
    isin_url = {}
    for cfg in FILES:
        if not cfg["path"].exists():
            continue
        wb = openpyxl.load_workbook(str(cfg["path"]), read_only=True)
        ws = wb[cfg["sheets"][0]]
        for r in range(cfg["data_start"], ws.max_row + 1):
            iv = ws.cell(r, cfg["col_isin"]).value
            iv = str(iv).strip() if iv else ''
            if not iv or iv in isin_url:
                continue
            u = fondidoc_url(ws.cell(r, cfg["col_fondidoc"]))
            if u:
                isin_url[iv] = u
        wb.close()
    return isin_url


def main():
    isin_url = collect_isin_urls()
    log(f"ISIN unici con link FondiDoc: {len(isin_url)} (scrape con {WORKERS} thread)")

    # scrape UNA volta sola, risultati per ISIN
    by_isin = {}
    done = errors = 0
    rev = {}
    for isin, u in isin_url.items():
        rev.setdefault(u, isin)  # un URL -> un ISIN
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = {ex.submit(scrape_fund, u): u for u in rev}
        for n, fut in enumerate(as_completed(futs), 1):
            u = futs[fut]
            try:
                by_isin[rev[u]] = fut.result()
                done += 1
            except Exception:
                errors += 1
            if n % 500 == 0:
                log(f"  {n}/{len(rev)} (ok {done}, err {errors})")
    log(f"Scraping finito: ok {done}, err {errors}")

    # applica a ciascun file secondo il suo schema
    for cfg in FILES:
        if not cfg["path"].exists():
            log(f"  SALTO {cfg['path'].name} (assente)")
            continue
        wb = openpyxl.load_workbook(str(cfg["path"]))
        for sheet in cfg["sheets"]:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            upd = 0
            for r in range(cfg["data_start"], ws.max_row + 1):
                iv = ws.cell(r, cfg["col_isin"]).value
                iv = str(iv).strip() if iv else ''
                d = by_isin.get(iv)
                if not d:
                    continue
                for col, key in cfg["perf"].items():
                    v = d.get(key)
                    if v is not None:
                        c = ws.cell(r, col)
                        c.value = v
                        c.number_format = '0.00%'
                upd += 1
            log(f"  {cfg['path'].name} / '{sheet}': {upd} righe aggiornate")
        wb.save(str(cfg["path"]))
    log("FINITO.")


if __name__ == "__main__":
    main()
